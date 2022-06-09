from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # Creating a new sheet with a name by default
ws.title = 'MySheet'  # Changing the name of the sheet
ws.sheet_properties.tabColor = 'ff66ff'

# Creating another sheet and giving it a name
ws1 = wb.create_sheet('YourSheet')
ws2 = wb.create_sheet('NewSheet', 2)  # Creating a new sheet at the index 2

new_ws = wb['NewSheet']  # Accessing 'NewSheet' in a form of dictionary

print(wb.sheetnames)  # Confirming the names of all the sheets

# Copying Sheets
new_ws['A1'] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'CopiedSheet'

wb.save('sample.xlsx')
