from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80:
        print("Student No.", row[0].value, " is eligible to apply for TOEFL.")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == 'English':
            cell.value == 'Computer'

wb.save("sample_modified.xlsx")
