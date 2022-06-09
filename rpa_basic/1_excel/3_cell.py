from random import *
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'MySheet'

# Entering a value '1' into the A1 cell
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 4
ws['B2'] = 5
ws['B3'] = 6

print(ws['A1'])  # Printing the 'info' of A1
print(ws['A1'].value)  # Printing the 'value' of A1
print(ws['A10'].value)  # Printing 'None' since there was no value entered

# row = 1, 2, 3, ...
# column = A, B, C ...
print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)

c = ws.cell(column=3, row=1, value=10)  # ws["C1"].value = 10
print(c.value)

index = 1
for x in range(1, 11):  # 10 rows
    for y in range(1, 11):  # 10 columns
        # ws.cell(row=x, column=y, value=randint(0, 100)) # any random numbers between 0~100
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save('sample.xlsx')
