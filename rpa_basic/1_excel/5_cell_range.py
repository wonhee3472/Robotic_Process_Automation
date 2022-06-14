from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# adding data by one row
# ID is Column A, English is Column B, and Math is Column C
ws.append(['ID', "English", "Math"])
for i in range(1, 11):  # Entering 10 data
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]  # saving Column B into a variable `col_B`

# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"]  # Saving both column A and B into a variable `col_range`
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]  # only the first row in the sheet
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]  # from the second row to the 6th row
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")  # A
#         print(xy[1], end=" ")  # 1
#     print()

# for row in ws.iter_rows(): # all rows in the worksheet
#     print(row[1].value)

# for column in ws.iter_cols():  # all columns in the worksheet
#     print(column[0].value)

# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")
